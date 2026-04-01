"""
Export service — Professional branded Excel and PDF/HTML reports.
Groups members by level, shows product sub-summaries per level,
and a master summary across the full project.
"""
import io
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.db import get_ref_db

# ── Nullifire Brand Constants ──
NF_RED = 'E6332A'
NF_NAVY = '1E3480'
DARK_GREY = '333333'
MID_GREY = '666666'
LIGHT_GREY = 'F2F2F7'
ALT_ROW = 'F8F8FB'
WHITE = 'FFFFFF'
STATUS_GREEN = '10B981'
STATUS_AMBER = 'F59E0B'
STATUS_RED = 'EF4444'

THIN_NAVY = Border(
    left=Side(style='thin', color=NF_NAVY), right=Side(style='thin', color=NF_NAVY),
    top=Side(style='thin', color=NF_NAVY), bottom=Side(style='thin', color=NF_NAVY),
)


def _resolve_names(members, products_map):
    """Add display names for fire_rating, failure_temp, product to each member."""
    db = get_ref_db()
    fr_cache = {}
    ft_cache = {}

    for m in members:
        # Fire rating
        frid = m.get('fire_rating_id')
        if frid and frid not in fr_cache:
            row = db.execute('SELECT description FROM fire_ratings WHERE id=?', (frid,)).fetchone()
            fr_cache[frid] = row['description'] if row else str(frid)
        m['fire_rating_name'] = fr_cache.get(frid, '')

        # Failure temp
        ftid = m.get('failure_temp_id')
        if ftid and ftid not in ft_cache:
            row = db.execute('SELECT description FROM failure_temps WHERE id=?', (ftid,)).fetchone()
            ft_cache[ftid] = row['description'] if row else str(ftid)
        m['failure_temp_name'] = ft_cache.get(ftid, '')

        # Product
        pid = m.get('product_id')
        m['product_name'] = products_map.get(pid, '') if pid else ''

    return members


def _group_by_level(members):
    """Group members by level. Returns ordered list of (level_name, members_list)."""
    groups = defaultdict(list)
    for m in members:
        level = m.get('level') or '(No Level)'
        groups[level].append(m)
    # Sort: named levels first alphabetically, '(No Level)' last
    keys = sorted(groups.keys(), key=lambda k: (k == '(No Level)', k))
    return [(k, groups[k]) for k in keys]


def _product_subtotals(members, products_map):
    """Calculate per-product subtotals for a group of members."""
    by_product = defaultdict(lambda: {'count': 0, 'area': 0, 'litres': 0, 'kg': 0})
    for m in members:
        pid = m.get('product_id') or 0
        name = products_map.get(pid, 'Project Default') if pid else 'Project Default'
        by_product[name]['count'] += 1
        by_product[name]['area'] += m.get('surface_area_m2', 0) or 0
        by_product[name]['litres'] += m.get('volume_litres', 0) or 0
        by_product[name]['kg'] += m.get('weight_kg', 0) or 0
    return dict(by_product)


# ═══════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════

def _xl_header_cell(cell):
    cell.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
    cell.fill = PatternFill(start_color=NF_NAVY, end_color=NF_NAVY, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_NAVY


def _xl_data_cell(cell, row_i, is_num=False):
    cell.font = Font(name='Calibri', size=10, color=DARK_GREY)
    bg = ALT_ROW if row_i % 2 == 0 else WHITE
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='right' if is_num else 'left', vertical='center')
    cell.border = Border(bottom=Side(style='hair', color='D0D0D0'))


def _xl_level_header(ws, row, label, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(name='Calibri', bold=True, size=12, color=NF_NAVY)
    cell.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')
    cell.border = Border(bottom=Side(style='medium', color=NF_RED))
    ws.row_dimensions[row].height = 22


def _xl_subtotal_row(ws, row, label, area, litres, kg, num_cols, col_map):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_map['area'] - 1)
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = Font(name='Calibri', bold=True, size=10, color=NF_NAVY)
    lbl.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')
    lbl.alignment = Alignment(horizontal='right')

    for c in range(2, col_map['area']):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')

    for c, v, fmt in [(col_map['area'], area, '0.00'), (col_map['litres'], litres, '0.00'), (col_map['kg'], kg, '0.0')]:
        cell = ws.cell(row=row, column=c, value=round(v, 2))
        cell.font = Font(name='Calibri', bold=True, size=10, color=NF_NAVY)
        cell.fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = fmt

    # Fill remaining columns
    for c in range(col_map['kg'] + 1, num_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type='solid')


def export_excel(project, members, product):
    """Export professional branded Excel with level grouping and product sub-summaries."""
    from app.services import product_service
    all_products = product_service.get_products()
    products_map = {p['id']: p['name'] for p in all_products}
    if product:
        products_map[0] = product['name']  # default product

    members = _resolve_names(list(members), products_map)
    levels = _group_by_level(members)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Specification'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    num_cols = 15
    HEADERS = [
        ('Section', 18), ('Type', 7), ('Exposure', 9), ('Hp/A', 7),
        ('Fire Rating', 13), ('Failure Temp', 12), ('Product', 15),
        ('DFT (mm)', 10), ('DFT (um)', 9),
        ('Qty', 5), ('Length (m)', 10),
        ('Area (m2)', 10), ('Litres', 8), ('Weight (kg)', 10),
        ('Status', 8),
    ]
    col_map = {'area': 12, 'litres': 13, 'kg': 14}

    # Set column widths
    for i, (_, w) in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title ──
    row = 1
    ws.merge_cells(f'A{row}:O{row}')
    c = ws.cell(row=row, column=1, value='NULLIFIRE INTUMESCENT SPECIFICATION')
    c.font = Font(name='Calibri', bold=True, size=18, color=NF_NAVY)
    ws.row_dimensions[row].height = 30

    # Red accent
    row = 2
    ws.merge_cells(f'A{row}:O{row}')
    ws.cell(row=row, column=1).fill = PatternFill(start_color=NF_RED, end_color=NF_RED, fill_type='solid')
    ws.row_dimensions[row].height = 4

    # ── Project Details ──
    product_name = product['name'] if product else 'N/A'
    details = [
        ('Project:', project.get('name', ''), 'Client:', project.get('client', '')),
        ('Product:', product_name, 'Date:', datetime.now().strftime('%d %B %Y')),
        ('Members:', str(len(members)), '', ''),
    ]
    row = 4
    for l1, v1, l2, v2 in details:
        for off, (la, va) in enumerate([(l1, v1), (l2, v2)]):
            if not la:
                continue
            lc = 1 + off * 5
            cl = ws.cell(row=row, column=lc, value=la)
            cl.font = Font(name='Calibri', bold=True, size=10, color=NF_NAVY)
            cl.alignment = Alignment(horizontal='right')
            cv = ws.cell(row=row, column=lc + 1, value=va)
            cv.font = Font(name='Calibri', size=10, color=DARK_GREY)
            ws.merge_cells(start_row=row, start_column=lc + 1, end_row=row, end_column=lc + 3)
        row += 1

    row += 1

    # ── Per-Level Sections ──
    for level_name, level_members in levels:
        # Level header
        _xl_level_header(ws, row, f'Level: {level_name}', num_cols)
        row += 1

        # Column headers
        for ci, (name, _) in enumerate(HEADERS, 1):
            _xl_header_cell(ws.cell(row=row, column=ci, value=name))
        ws.row_dimensions[row].height = 22
        row += 1

        # Data rows
        for di, m in enumerate(level_members):
            dft = m.get('dft_mm')
            vals = [
                (m.get('section_name', ''), False),
                (m.get('steel_type', ''), False),
                (m.get('hp_profile_name', ''), False),
                (round(m['hp_over_a']) if m.get('hp_over_a') else None, True),
                (m.get('fire_rating_name', ''), False),
                (m.get('failure_temp_name', ''), False),
                (m.get('product_name', '') or product_name, False),
                (dft, True),
                (round(dft * 1000) if dft else None, True),
                (m.get('quantity', 1), True),
                (m.get('length_m', 0), True),
                (m.get('surface_area_m2', 0), True),
                (m.get('volume_litres', 0), True),
                (m.get('weight_kg', 0), True),
                ((m.get('status', '') or '').upper(), False),
            ]
            for ci, (v, is_n) in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                _xl_data_cell(cell, di, is_n)
                if ci == 8 and v:
                    cell.number_format = '0.000'
                elif ci in (11, 12) and isinstance(v, (int, float)):
                    cell.number_format = '0.00'
                elif ci in (13, 14) and isinstance(v, (int, float)):
                    cell.number_format = '0.00'

            # Color status
            sc = ws.cell(row=row, column=15)
            cm = {'OK': STATUS_GREEN, 'WARNING': STATUS_AMBER, 'EXCEEDS': STATUS_RED}
            sv = (m.get('status', '') or '').upper()
            if sv in cm:
                sc.font = Font(name='Calibri', bold=True, size=10, color=cm[sv])
            row += 1

        # Product sub-summaries for this level
        psubs = _product_subtotals(level_members, products_map)
        if len(psubs) > 1 or (len(psubs) == 1 and 'Project Default' not in psubs):
            for pname, vals in sorted(psubs.items()):
                _xl_subtotal_row(ws, row, f'{pname} ({vals["count"]} members)',
                                 vals['area'], vals['litres'], vals['kg'], num_cols, col_map)
                row += 1

        # Level totals
        lv_area = sum(m.get('surface_area_m2', 0) or 0 for m in level_members)
        lv_litres = sum(m.get('volume_litres', 0) or 0 for m in level_members)
        lv_kg = sum(m.get('weight_kg', 0) or 0 for m in level_members)
        _xl_subtotal_row(ws, row, f'Level Total: {level_name} ({len(level_members)} members)',
                         lv_area, lv_litres, lv_kg, num_cols, col_map)
        row += 2

    # ── Master Summary ──
    _xl_level_header(ws, row, 'PROJECT SUMMARY', num_cols)
    row += 1

    # Per-product master totals
    master_psubs = _product_subtotals(members, products_map)
    for pname, vals in sorted(master_psubs.items()):
        pn = pname if pname != 'Project Default' else product_name
        _xl_subtotal_row(ws, row, f'{pn} ({vals["count"]} members)',
                         vals['area'], vals['litres'], vals['kg'], num_cols, col_map)
        row += 1

    # Grand total
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_map['area'] - 1)
    gt_lbl = ws.cell(row=row, column=1, value='GRAND TOTAL')
    gt_lbl.font = Font(name='Calibri', bold=True, size=12, color=WHITE)
    gt_lbl.fill = PatternFill(start_color=NF_NAVY, end_color=NF_NAVY, fill_type='solid')
    gt_lbl.alignment = Alignment(horizontal='right')
    for c in range(2, col_map['area']):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=NF_NAVY, end_color=NF_NAVY, fill_type='solid')

    t_area = sum(m.get('surface_area_m2', 0) or 0 for m in members)
    t_litres = sum(m.get('volume_litres', 0) or 0 for m in members)
    t_kg = sum(m.get('weight_kg', 0) or 0 for m in members)
    for c, v, fmt in [(col_map['area'], t_area, '0.00'), (col_map['litres'], t_litres, '0.00'), (col_map['kg'], t_kg, '0.0')]:
        cell = ws.cell(row=row, column=c, value=round(v, 2))
        cell.font = Font(name='Calibri', bold=True, size=12, color=WHITE)
        cell.fill = PatternFill(start_color=NF_NAVY, end_color=NF_NAVY, fill_type='solid')
        cell.alignment = Alignment(horizontal='right')
        cell.number_format = fmt
    for c in range(col_map['kg'] + 1, num_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(start_color=NF_NAVY, end_color=NF_NAVY, fill_type='solid')

    # Footer
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    f = ws.cell(row=row, column=1,
                value=f'Tremco CPG UK Limited  |  Generated {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Nullifire Intumescent Calculator')
    f.font = Font(name='Calibri', size=8, color=MID_GREY, italic=True)
    f.alignment = Alignment(horizontal='center')

    ws.freeze_panes = 'A8'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════
#  PDF / HTML EXPORT
# ═══════════════════════════════════════════

def export_pdf_html(project, members, product, summary):
    from app.services import product_service
    all_products = product_service.get_products()
    products_map = {p['id']: p['name'] for p in all_products}
    product_name = product['name'] if product else 'N/A'
    if product:
        products_map[0] = product['name']

    members = _resolve_names(list(members), products_map)
    levels = _group_by_level(members)

    # Build level sections
    sections_html = ''
    for level_name, level_members in levels:
        rows = ''
        for i, m in enumerate(level_members):
            dft = m.get('dft_mm')
            dft_str = f'{dft:.3f} ({int(dft*1000)} &micro;m)' if dft else '&mdash;'
            status = m.get('status', '')
            sc = {'ok': '#10b981', 'warning': '#f59e0b', 'exceeds': '#ef4444'}.get(status, '#9ca3af')
            bg = '#f8f8fb' if i % 2 == 0 else '#fff'
            pname = m.get('product_name', '') or product_name
            rows += f'''<tr style="background:{bg}">
                <td>{m.get('section_name','')}</td><td>{m.get('steel_type','')}</td>
                <td>{m.get('hp_profile_name','')}</td><td class="r">{int(m.get('hp_over_a',0) or 0)}</td>
                <td>{m.get('fire_rating_name','')}</td><td>{m.get('failure_temp_name','')}</td>
                <td>{pname}</td><td class="r">{dft_str}</td>
                <td class="r">{m.get('quantity',1)}</td><td class="r">{m.get('length_m',0):.2f}</td>
                <td class="r">{m.get('surface_area_m2',0):.2f}</td><td class="r">{m.get('volume_litres',0):.2f}</td>
                <td class="r">{m.get('weight_kg',0):.1f}</td>
                <td style="text-align:center"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{sc}"></span></td>
            </tr>'''

        # Product sub-summaries
        psubs = _product_subtotals(level_members, products_map)
        sub_html = ''
        if len(psubs) > 1 or (len(psubs) == 1 and 'Project Default' not in psubs):
            for pn, v in sorted(psubs.items()):
                sub_html += f'<tr class="sub"><td colspan="10" style="text-align:right">{pn} ({v["count"]} members)</td><td class="r">{v["area"]:.2f}</td><td class="r">{v["litres"]:.2f}</td><td class="r">{v["kg"]:.1f}</td><td></td></tr>'

        lv_a = sum(m.get('surface_area_m2', 0) or 0 for m in level_members)
        lv_l = sum(m.get('volume_litres', 0) or 0 for m in level_members)
        lv_k = sum(m.get('weight_kg', 0) or 0 for m in level_members)

        sections_html += f'''
        <div class="level-section">
            <div class="level-header">{level_name}</div>
            <table><thead><tr>
                <th>Section</th><th>Type</th><th>Exp.</th><th class="r">Hp/A</th>
                <th>Fire Rating</th><th>Temp</th><th>Product</th><th class="r">DFT</th>
                <th class="r">Qty</th><th class="r">Len(m)</th><th class="r">Area</th>
                <th class="r">Litres</th><th class="r">Wt(kg)</th><th></th>
            </tr></thead><tbody>{rows}{sub_html}
            <tr class="level-total"><td colspan="10" style="text-align:right;font-weight:700">Level Total ({len(level_members)} members)</td>
                <td class="r" style="font-weight:700">{lv_a:.2f}</td><td class="r" style="font-weight:700">{lv_l:.2f}</td>
                <td class="r" style="font-weight:700">{lv_k:.1f}</td><td></td></tr>
            </tbody></table>
        </div>'''

    # Master summary
    master_psubs = _product_subtotals(members, products_map)
    master_rows = ''
    for pn, v in sorted(master_psubs.items()):
        pn_display = pn if pn != 'Project Default' else product_name
        master_rows += f'<div class="ms-row"><span class="ms-name">{pn_display}</span><span class="ms-detail">{v["count"]} members &nbsp;|&nbsp; {v["area"]:.1f} m&sup2; &nbsp;|&nbsp; <b>{v["litres"]:.1f} L</b> &nbsp;|&nbsp; {v["kg"]:.1f} kg</span></div>'

    t_a = sum(m.get('surface_area_m2', 0) or 0 for m in members)
    t_l = sum(m.get('volume_litres', 0) or 0 for m in members)
    t_k = sum(m.get('weight_kg', 0) or 0 for m in members)

    return f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<title>Nullifire Specification &mdash; {project["name"]}</title>
<style>
    @page {{ size: A4 landscape; margin: 12mm; }}
    body {{ font-family: Calibri, 'Segoe UI', Arial, sans-serif; font-size: 9px; color: #333; margin: 0; padding: 16px; }}
    .hdr {{ background: #1E3480; color: #fff; padding: 14px 20px; margin: -16px -16px 0; display: flex; align-items: center; justify-content: space-between; }}
    .hdr h1 {{ font-size: 16px; font-weight: 700; letter-spacing: 1px; margin: 0; }}
    .hdr .sub {{ font-size: 10px; opacity: 0.6; }}
    .hdr-r {{ text-align: right; font-size: 9px; }}
    .hdr-r b {{ display: block; font-size: 10px; }}
    .accent {{ height: 3px; background: #E6332A; margin: 0 -16px 12px; }}
    .meta {{ display: flex; gap: 30px; margin-bottom: 10px; }}
    .meta-item .ml {{ font-size: 8px; color: #666; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 700; }}
    .meta-item .mv {{ font-size: 12px; font-weight: 600; color: #1E3480; }}
    .level-section {{ margin-bottom: 14px; }}
    .level-header {{ font-size: 12px; font-weight: 700; color: #1E3480; padding: 5px 0; border-bottom: 2px solid #E6332A; margin-bottom: 4px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 9px; }}
    th {{ background: #1E3480; color: #fff; padding: 4px 5px; text-align: left; font-size: 8px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.2px; }}
    td {{ padding: 3px 5px; border-bottom: 1px solid #e8e8ee; }}
    .r {{ text-align: right; }}
    .sub td {{ background: #f2f2f7; font-size: 9px; color: #1E3480; font-weight: 600; }}
    .level-total td {{ background: #f2f2f7; border-top: 2px solid #1E3480; color: #1E3480; }}
    .master {{ background: #f2f2f7; border: 2px solid #1E3480; border-radius: 6px; padding: 12px 16px; margin-top: 8px; }}
    .master h3 {{ font-size: 12px; color: #1E3480; margin-bottom: 8px; border-bottom: 2px solid #E6332A; padding-bottom: 4px; }}
    .ms-row {{ display: flex; justify-content: space-between; padding: 4px 0; border-bottom: 1px solid #ddd; font-size: 10px; }}
    .ms-name {{ font-weight: 700; color: #1E3480; }}
    .ms-detail {{ color: #666; }}
    .ms-detail b {{ color: #E6332A; }}
    .grand {{ display: flex; gap: 30px; margin-top: 10px; padding-top: 8px; border-top: 3px solid #1E3480; }}
    .grand-item .gl {{ font-size: 8px; color: #666; text-transform: uppercase; font-weight: 700; }}
    .grand-item .gv {{ font-size: 18px; font-weight: 800; color: #E6332A; }}
    .footer {{ margin-top: 14px; font-size: 7px; color: #999; border-top: 2px solid #E6332A; padding-top: 6px; text-align: center; }}
    .footer b {{ color: #1E3480; }}
    @media print {{ body {{ margin: 0; padding: 8px; }} .hdr {{ margin: -8px -8px 0; }} .accent {{ margin: 0 -8px 8px; }} }}
</style></head><body>
<div class="hdr"><div><h1>NULLIFIRE INTUMESCENT SPECIFICATION</h1><div class="sub">Structural Steel Fire Protection</div></div>
<div class="hdr-r"><b>Tremco CPG UK Limited</b><span style="opacity:0.6">protect@nullifire.com | www.nullifire.com</span></div></div>
<div class="accent"></div>
<div class="meta">
    <div class="meta-item"><div class="ml">Project</div><div class="mv">{project["name"]}</div></div>
    <div class="meta-item"><div class="ml">Client</div><div class="mv">{project.get("client","&mdash;")}</div></div>
    <div class="meta-item"><div class="ml">Default Product</div><div class="mv">{product_name}</div></div>
    <div class="meta-item"><div class="ml">Date</div><div class="mv">{datetime.now().strftime("%d %B %Y")}</div></div>
    <div class="meta-item"><div class="ml">Members</div><div class="mv">{len(members)}</div></div>
</div>
{sections_html}
<div class="master">
    <h3>PROJECT SUMMARY</h3>
    {master_rows}
    <div class="grand">
        <div class="grand-item"><div class="gl">Total Area</div><div class="gv">{t_a:.1f} m&sup2;</div></div>
        <div class="grand-item"><div class="gl">Total Litres</div><div class="gv">{t_l:.1f} L</div></div>
        <div class="grand-item"><div class="gl">Total Weight</div><div class="gv">{t_k:.1f} kg</div></div>
        <div class="grand-item"><div class="gl">Total Members</div><div class="gv">{len(members)}</div></div>
    </div>
</div>
<div class="footer"><b>Tremco CPG UK Limited</b> &nbsp;|&nbsp; Torrington Avenue, Coventry, CV4 9TJ &nbsp;|&nbsp; protect@nullifire.com &nbsp;|&nbsp; www.nullifire.com<br>
Generated {datetime.now().strftime("%d/%m/%Y %H:%M")} &nbsp;|&nbsp; Nullifire Intumescent Calculator</div>
</body></html>'''
